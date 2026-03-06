import json, os, time, requests
from datetime import datetime
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

TOKEN_FILE = r"C:\Users\cesar.zarovski\bling_api\bling_tokens.json"
OUT = r"C:\Users\cesar.zarovski\Documents\DRE_2025_Bling_NFe.xlsx"
CACHE = r"C:\Users\cesar.zarovski\bling_api\nfe_2025_cache.jsonl"

with open(TOKEN_FILE, "r", encoding="utf-8") as f:
    tokens = json.load(f)
access_token = tokens.get("access_token")

headers = {
    "Authorization": f"Bearer {access_token}",
    "Content-Type": "application/json",
    "enable-jwt": "1",
}

start_date = "2025-01-01"
end_date = "2025-12-31"

base_url = "https://api.bling.com.br/Api/v3/nfe"
page = 1
limit = 100

seen = set()
if os.path.exists(CACHE):
    with open(CACHE, "r", encoding="utf-8") as f:
        for line in f:
            try:
                obj = json.loads(line)
                seen.add(obj.get("id"))
            except:
                pass

new_count = 0
while True:
    params = {
        "pagina": page,
        "limite": limit,
        "dataEmissaoInicial": start_date,
        "dataEmissaoFinal": end_date,
    }
    resp = requests.get(base_url, headers=headers, params=params, timeout=30)
    if resp.status_code != 200:
        raise SystemExit(f"List error {resp.status_code}: {resp.text}")
    data = resp.json().get("data", [])
    if not data:
        break
    for item in data:
        id_ = item.get("id")
        if not id_ or id_ in seen:
            continue
        durl = f"{base_url}/{id_}"
        dresp = requests.get(durl, headers=headers, timeout=30)
        if dresp.status_code != 200:
            continue
        det = dresp.json().get("data", {})
        val = det.get("valorNota") or 0
        data_emissao = det.get("dataEmissao") or det.get("dataOperacao")
        if data_emissao:
            dt = datetime.strptime(data_emissao.split()[0], "%Y-%m-%d")
            rec = {"id": id_, "dataEmissao": dt.strftime("%Y-%m-%d"), "valorNota": float(val)}
            with open(CACHE, "a", encoding="utf-8") as f:
                f.write(json.dumps(rec) + "\n")
            seen.add(id_)
            new_count += 1
        time.sleep(0.1)
    page += 1

# Build DRE from cache
monthly = defaultdict(float)
raw_rows = []
if os.path.exists(CACHE):
    with open(CACHE, "r", encoding="utf-8") as f:
        for line in f:
            try:
                obj = json.loads(line)
            except:
                continue
            dt = datetime.strptime(obj["dataEmissao"], "%Y-%m-%d")
            key = dt.strftime("%Y-%m")
            monthly[key] += obj.get("valorNota", 0)
            raw_rows.append(obj)

months = ["Jan","Fev","Mar","Abr","Mai","Jun","Jul","Ago","Set","Out","Nov","Dez","Total"]
wb = Workbook()
ws = wb.active
ws.title = "DRE_2025"

ws.cell(row=1, column=1, value="Conta").font = Font(bold=True)
for i, m in enumerate(months, start=2):
    ws.cell(row=1, column=i, value=m).font = Font(bold=True)
    ws.cell(row=1, column=i).alignment = Alignment(horizontal="center")

rows = [
    "Receita Bruta (NFe)",
    "(-) Impostos e Deduções",
    "Receita Líquida",
    "(-) CMV",
    "Lucro Bruto",
    "(-) Despesas Operacionais",
    "EBITDA",
    "(-) Depreciação e Amortização",
    "Resultado Operacional",
    "(+/-) Resultado Financeiro",
    "Lucro Líquido",
]

for r, name in enumerate(rows, start=2):
    ws.cell(row=r, column=1, value=name)

for i, m in enumerate(range(1,13), start=2):
    key = f"2025-{m:02d}"
    ws.cell(row=2, column=i, value=round(monthly.get(key, 0.0), 2))

for col in range(2, 15):
    c = get_column_letter(col)
    ws.cell(row=4, column=col, value=f"={c}2-{c}3")
    ws.cell(row=6, column=col, value=f"={c}4-{c}5")
    ws.cell(row=8, column=col, value=f"={c}6-{c}7")
    ws.cell(row=10, column=col, value=f"={c}8-{c}9")
    ws.cell(row=12, column=col, value=f"={c}10+{c}11")

for r in range(2, 13):
    ws.cell(row=r, column=14, value=f"=SUM(B{r}:M{r})")

ws.column_dimensions["A"].width = 32
for col in range(2, 15):
    ws.column_dimensions[get_column_letter(col)].width = 14
    for r in range(2, 13):
        ws.cell(row=r, column=col).number_format = 'R$ #,##0.00'

ws2 = wb.create_sheet("NFe_2025")
ws2.append(["id","dataEmissao","valorNota"])
for row in raw_rows:
    ws2.append([row["id"], row["dataEmissao"], row["valorNota"]])
ws2.column_dimensions["A"].width = 16
ws2.column_dimensions["B"].width = 12
ws2.column_dimensions["C"].width = 14
for r in range(2, ws2.max_row+1):
    ws2.cell(r, 3).number_format = 'R$ #,##0.00'

wb.save(OUT)
print(f"New fetched: {new_count}")
print(OUT)
